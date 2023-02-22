import calendar
from datetime import timedelta, datetime
import pandas as pd

from fastapi.responses import FileResponse
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from common_func import str2datetime, datetime2str, start, first_day, step
from const import NO_CHECKOUT, STATUS_OK, IGNORE_EMP_ID, STATUS_DAY_OFF

out_wb = load_workbook("WorkingdayTemplate.xlsx")
out_ws = out_wb.active


def handle_get_data_from_file(start: datetime, end: datetime):
    k = 0
    working_days = []
    step = timedelta(days=1)
    while start <= end:
        column = 6 + k
        out_ws.cell(row=6, column=column, value=start.strftime("%a"))
        out_ws.cell(row=7, column=column, value=start.day)
        working_days.append(start)
        k += 1
        start += step
    return working_days


def append_off_day(list_days: list, list_wking_days: list):
    list_days_emp = [str(day_emp.get("day"))[0:10] for day_emp in list_days]
    for wking in list_wking_days:
        if datetime2str(wking) in list_days_emp:
            continue
        item = {
            "day": wking,
            "late_time": 1,
            "checkout_time": "",
            "time_str": "",
        }
        list_days.append(item)
    return list_days


def get_late_time(time: float) -> float:
    if time >= 8:
        return 8
    return time


def handle_open_file(excel_file, working_days, data):
    df = pd.read_excel(excel_file, skiprows=[0])
    for row in (df.values.tolist()):
        emp_id = row[0].upper()
        emp_name = row[1].upper()
        key = (emp_id, emp_name)
        if key not in data:
            data[key] = []
        day = str2datetime(row[2])
        checkout_time = str2datetime(row[3]) if row[3] else NO_CHECKOUT
        right_time_str = f"{str(row[2])[0:10]} 08:30:00"
        noon_time_str = f"{str(row[2])[0:10]} 13:00:00"
        right_co_time_str = f"{str(row[2])[0:10]} 17:30:00"
        right_co_time = str2datetime(right_co_time_str)
        late_time_raw = 0
        late_time_real = 0

        if str2datetime(row[2]) > str2datetime(right_time_str):
            late_time_raw = str2datetime(row[2]) - str2datetime(right_time_str)

        if late_time_raw:
            late_time_real = late_time_raw.seconds / 3600
            if late_time_raw.seconds > 16200:
                late_time_raw = (
                        str2datetime(row[2])
                        - str2datetime(right_time_str)
                        + timedelta(hours=4)
                ).seconds

        late_time = get_late_time(late_time_real)

        if checkout_time != NO_CHECKOUT and checkout_time < str2datetime(
                right_co_time_str
        ):
            late_time += (right_co_time - checkout_time).seconds / 3600
        if (row[3] and row[2]) and (
                str2datetime(row[3]) - timedelta(minutes=1) == str2datetime(row[2])
        ):
            late_time = 8

        data_key = {
            "day": day,
            "late_time": get_late_time(late_time),
            "checkout_time": checkout_time,
            "time_str": str(row[2])[0:10],
        }
        data[key].append(data_key)

    for emp_id, emp_name in data:
        key = (emp_id, emp_name)
        data[key] = append_off_day(list_days=data[key], list_wking_days=working_days)
    return data


def get_export_info():
    now = datetime.today()
    month_num = datetime.now().month
    year = datetime.now().year
    prev_month_num = datetime.today().replace(month=month_num - 1).month if now.month != 1 else (12, year - 1)
    month = datetime.now().strftime("%b")
    p_day = now.replace(day=21, month=prev_month_num, year=year)
    export_day = str(p_day)[0:10]
    return month, export_day


def handle_export_file(data):
    month, export_day = get_export_info()
    days_in_month = calendar.monthrange(start.year, start.month)[1]
    plus_day = timedelta(days=days_in_month - 1)
    end = start + plus_day
    k = 0  # column
    start_day = datetime.strptime(export_day, "%Y-%m-%d")
    time_now = f"{month}_{datetime.now().year}"
    out_ws.cell(row=2, column=1).value = f"TIME ATTENDANCE RECORD {time_now}"
    from_day = export_day.replace("-", "/")
    end_day = str(end)[0:10].replace("-", "/")
    out_ws.cell(row=4, column=1).value = f"WORKING DAY {from_day} - {end_day}"
    while start_day <= end:
        column = 6 + k
        out_ws.cell(row=6, column=column, value=start_day.strftime("%a"))
        out_ws.cell(row=7, column=column, value=start_day.day)

        i = 0  # row
        for emp_id, emp_name in sorted(data):
            row = 8 + i
            out_ws.cell(row=row, column=1).value = i + 1
            out_ws.cell(row=row, column=2).value = emp_id
            out_ws.cell(row=row, column=3).value = emp_name

            check_time = sorted(data[(emp_id, emp_name)], key=lambda c: c["day"])
            for woking in check_time:
                try:
                    day = woking.get("day")
                    late_time = woking.get("late_time")
                    checkout_time = woking.get("checkout_time")
                    time_str = woking.get("time_str")
                    # Column start from 6 because there 2 hidden columns
                    comments = []
                    comment_tool = ""
                    cell = out_ws.cell(row=row, column=column)
                    time = day.strftime("%H:%M:%S")
                    status = STATUS_OK

                    co_time_str = f"{str(day)[0:10]} 17:30:00"
                    co_time = str2datetime(co_time_str)
                    comments.append(f"Checkout lúc: {checkout_time}")
                    if emp_id in IGNORE_EMP_ID:
                        status = STATUS_OK
                        comments = []
                    elif start_day.weekday() > 4:
                        status = STATUS_OK
                    elif start_day == day:
                        cell.fill = PatternFill(patternType="solid", fgColor="FCBA03")
                        status = STATUS_DAY_OFF
                        comments.append("Không checkin")
                    elif day.strftime("%m%d") != start_day.strftime("%m%d"):
                        comments = []
                        continue
                    elif late_time > 0:
                        status = late_time * 0.125
                        comments.insert(0, f"Checkin lúc: {time}")
                    elif checkout_time == NO_CHECKOUT:
                        status = STATUS_DAY_OFF
                        cell.fill = PatternFill(patternType="solid", fgColor="ffffff")
                    else:
                        status = STATUS_OK
                        cell.fill = PatternFill(patternType="solid", fgColor="ffffff")
                    if status != STATUS_OK:
                        cell.value = status
                        print(f"---{emp_name}---{cell} --- {status}")
                    if status == STATUS_DAY_OFF:
                        cell.fill = PatternFill(patternType="solid", fgColor="FCBA03")
                    if comments:
                        comments.insert(0, f"TDT STAFF: ")
                        comment_tool = "\n".join(comments)
                        cell.comment = Comment(comment_tool, "Tool")
                    if status == STATUS_OK:
                        cell.comment = None
                except Exception as e:
                    print(f"{woking} sida, {e}")
            sum_cell = out_ws.cell(row=row, column=38)
            sum_cell.value = f"=SUM(F{row}:AJ{row})"

            i += 1

        k += 1
        start_day += step

    OUT_FILE = f"TimeSheet.xlsx"
    out_wb.save(OUT_FILE)
    print("Exported")

